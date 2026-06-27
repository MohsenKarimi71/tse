from django.shortcuts import render, redirect
from django.conf import settings
from django.http import JsonResponse
import os
import jdatetime
import datetime
import requests
from openpyxl import load_workbook

from tse_downloader.models import Stock, Option, OptionPrice, StockPrice
from tse_downloader.tools import parse_option_description
from file_tracker.models import ExcelFileInfo


# DOWNLOAD_LINK = "https://members.tsetmc.com/tsev2/excel/MarketWatchPlus.aspx?d=0&format=0"
# DOWNLOAD_LINK = "https://members.tsetmc.com/tsev2/excel/MarketWatchPlus.aspx?d=0"
DOWNLOAD_LINK = "https://old.tsetmc.com/tsev2/excel/MarketWatchPlus.aspx?d=0"
EXCEL_FILES_SAVE_PATH = os.path.join(settings.BASE_DIR, 'excel_files')


def setup_download_tse_excel_data(request):
    return render(request, "tse_downloader/setup.html")   


def download_tse_excel_data(request):
    """
    Download the latest excel file from TSETMC,
    save it on disk and return the saved file's path.
    """
    response = requests.get(DOWNLOAD_LINK)

    daily_directory_path = os.path.join(EXCEL_FILES_SAVE_PATH, jdatetime.date.today().strftime("%Y%m%d"))
    os.makedirs(daily_directory_path, exist_ok=True)

    timestamp = jdatetime.datetime.now().time().strftime("%H%M%S")
    path = os.path.join(daily_directory_path, timestamp + '.xlsx')
    
    output = open(path, 'wb')
    output.write(response.content)
    output.close()
    excel_file = ExcelFileInfo(directory=jdatetime.date.today().strftime("%Y%m%d"), name=timestamp + '.xlsx')
    excel_file.save()

    # return render(request, "tse_downloader/setup.html") 
    return JsonResponse({"msg":excel_file.name})


def save_stocks_with_options_2db_from_file(request, directory, name):
    """
    این ویو سعی می کند از اطلاعات یک فایل تمامی سهامی را که
    دارای اختیار معامله می باشند پیدا کرده و در پایگاه داده ذخیره نماید
    """
    excel_file_path = os.path.join(settings.BASE_DIR,
        "excel_files",
        directory,
        name
    )
    print(excel_file_path)

    wb = load_workbook(excel_file_path, read_only=True, data_only=True)
    sheet = wb.active

    # finding excel file last row
    last_row = 3  # starting from row 4
    for row in sheet.iter_rows(min_row=4, min_col=1, max_col=1, values_only=True):
        if row[0] is None:
            break
        last_row += 1

    stored_option_chain_stock = list(Stock.objects.filter(is_in_option_chain=True))

    option_stocks = len(stored_option_chain_stock)
    print("currently stored_option_chain_stock option chain: ", option_stocks)

    for row in sheet.iter_rows(min_row=4, max_row=last_row, min_col=1, max_col=2):
        text = row[1].value
        if('اختيار' in text):
            is_stored = False
            try:
                option_type, symbol, strike_price, expiry_date = parse_option_description(text)
            except Exception as e:
                break
            for stock in stored_option_chain_stock:
                if(symbol == stock.symbol):
                    is_stored = True
                    break

            # find the symbol and it's name
            if not is_stored:
                is_symbol_found = False
                for row in sheet.iter_rows(min_row=4, max_row=last_row, min_col=1, max_col=2):
                    if row[0].value == symbol:
                        name = row[1].value
                        is_symbol_found = True
                        break

                # save the stock into database
                if is_symbol_found:
                    new_option_chain_stock = Stock(
                        symbol=symbol,
                        name=name,
                        group="OTHER",
                        is_in_option_chain=True
                    )
                    new_option_chain_stock.save()
                    option_stocks += 1
                    print("new option chain:", symbol)
                    stored_option_chain_stock.append(new_option_chain_stock)
                else:
                    print("No symbol found for symbol >>> ", symbol, " <<<")
                    
    print("final stored_option_chain_stock option chain: ", option_stocks)

    return redirect("file_tracker:file_list")


def save_new_option_contracts_2db_from_file(request, directory, name):
    """
    در این ویو قراردادهای اختیار معامله موجود در فایل مورد نظر
    پیدا و در صورتی که تاریخ سررسید آنها نرسیده باشد و قبلا به لیست
    قراردادها اضافه نشده باشد آنها را به پایگاه داده اضافه می کند
    """
    excel_file_path = os.path.join(settings.BASE_DIR,
        "excel_files",
        directory,
        name
    )
    print(excel_file_path)
    wb = load_workbook(excel_file_path)
    sheet = wb.active
    
    # finding excel file last row
    last_row = 3  # starting from row 4
    for row in sheet.iter_rows(min_row=4, min_col=1, max_col=1, values_only=True):
        if row[0] is None:
            break
        last_row += 1

    for row in sheet.iter_rows(min_row=4, min_col=1, max_col=21, max_row=last_row):
        text = row[1].value
        if('اختيار' in text):
            
            result = parse_option_description(text)
            if result:
                option_type, symbol, strike_price, expiry_date = result
                try:
                    stock = Stock.objects.get(symbol=symbol)

                    # save the Option contract into database if contract is not expired!
                    if expiry_date >= datetime.date.today():
                        contract_data = {
                            "stock": stock,
                            "strike_price": strike_price,
                            "expiration_date": expiry_date,
                            "contract_type": option_type
                        }

                        # Try to get the contract or create it if not exists
                        option_contract = Option.objects.get_or_create(**contract_data)
                except Exception as e:
                    print(str(e), "Error in saving contracts >>> ", symbol, " <<<")
                    
        
    file = ExcelFileInfo.objects.get(directory=directory, name=name)
    file.is_contracts_saved = True
    file.save()
    return redirect("file_tracker:file_list")


def save_option_contracts_deal_info_2db_from_file(request, directory, name):
    """
    در این ویو آخرین اطلاعات مربوط به معامله اختیارهای معامله
    از فایل مورد نظر استخراج و در پایگاه داده ذخیره می گردد
    """
    excel_file_path = os.path.join(settings.BASE_DIR,
        "excel_files",
        directory,
        name
    )
    print(excel_file_path)
    wb = load_workbook(excel_file_path)
    sheet = wb.active
    
    # finding excel file last row
    last_row = 3  # starting from row 4
    for row in sheet.iter_rows(min_row=4, min_col=1, max_col=1, values_only=True):
        if row[0] is None:
            break
        last_row += 1

    # saving option contracrs prices
    for row in sheet.iter_rows(min_row=4, min_col=1, max_col=21, max_row=last_row):
        text = row[1].value
        if('اختيار' in text):
            
            result = parse_option_description(text)
            if result:
                option_type, symbol, strike_price, expiry_date = result
                try:
                    stock = Stock.objects.get(symbol=symbol)

                    # save the Option contract into database if contract is not expired!
                    if expiry_date >= datetime.date.today():
                        # check if any deal is occured
                        if(int(row[2].value) > 0):
                            contract_data = {
                                "stock": stock,
                                "strike_price": strike_price,
                                "expiration_date": expiry_date,
                                "contract_type": option_type
                            }

                            # Try to get the contract or create it if not exists
                            option_contract = Option.objects.get_or_create(**contract_data)
                            # first delete other option's prices
                            deal_count = row[2].value
                            deal_volume = row[3].value
                            deal_value = int(deal_volume) * int(row[10].value) # row[10] ==> final price
                            # هر یک واحد از حجم قراردادهای اختیار برابر با هزار سهم است. برای
                            # این موضوع ارزش معاملات باید در هزار ضرب شود
                            # و برای تبدیل به تومان تقسیم بر 10 بشود 
                            deal_value *= 100
                            last_deal_price = row[7].value
                            buy_bid_price = row[19].value
                            sell_bid_price = row[20].value

                            option_price_obj = OptionPrice(
                                option = option_contract[0],
                                deal_count = deal_count,
                                deal_volume = deal_volume,
                                deal_value = deal_value,
                                last_deal_price = last_deal_price,
                                buy_bid_price = buy_bid_price,
                                sell_bid_price = sell_bid_price
                            )

                            option_price_obj.save()

                except Exception as e:
                    print(str(e), "Error in saving deal info >>> ", symbol, " <<<")

    # saving option chain stocks prices
    option_chain = Stock.objects.filter(is_in_option_chain=True)
    for stock in option_chain:
        for row in sheet.iter_rows(min_row=4, min_col=0, max_col=21, max_row=last_row):
            text = row[0].value
            if(text == stock.symbol):  
                
                stock_price_obj = StockPrice(
                    stock = stock,
                    price = row[20].value
                )
                stock_price_obj.save()
                break

    file = ExcelFileInfo.objects.get(directory=directory, name=name)
    file.is_prices_saved = True
    file.save()
    return redirect("file_tracker:file_list")


def delete_all_stocks_with_options(request):
    """این ویو تمامی سهامی را که در زنجیره اختیار معامله قرار دارند٬ حذف می کند"""
    Stock.objects.filter(is_in_option_chain=True).delete()
    return redirect("option_visualizer:option_chain")


def delete_stock_with_options(request):
    pass


def delete_all_contracts(request):
    Option.objects.all().delete()
    return redirect("option_visualizer:option_chain")


def delete_all_expired_contracts(request):
    Option.objects.filter(expiration_date__lte=datetime.date.today()).delete()
    return redirect("option_visualizer:all_options_contracts")


def delete_stock_contracts(request, stock_id):
    stock = Stock.objects.get(pk=stock_id)
    stock.contracts.all().delete()
    return redirect("option_visualizer:option_chain")


def delete_stock_expired_contracts(request, stock_id):
    stock = Stock.objects.get(pk=stock_id)
    stock.contracts.filter(expiration_date__lt=datetime.date.today()).delete()
    return redirect("option_visualizer:option_contracts", pk=stock_id)


def delete_all_stocks_contracts_deal_info(request):
    OptionPrice.objects.all().delete()
    return redirect("option_visualizer:all_options_contracts")


def delete_stock_contracts_deal_info(request, stock_id):
    stock = Stock.objects.get(pk=stock_id)
    contracts = stock.contracts.all()
    for contract in contracts:
        OptionPrice.objects.filter(option=contract.id).delete()
    return redirect("option_visualizer:option_contracts", pk=stock_id)