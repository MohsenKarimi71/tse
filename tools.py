### This file contains functions to download and work on Iran's stock market data files.
import os
import datetime
import sqlite3
import requests
import re

from openpyxl import load_workbook

from variables import DOWNLOAD_LINK 


def download_excel_data_file(daily_path):
    """Download the latest excel file from TSETMC, save it on disk and return the saved file's path."""
    path = os.path.join(daily_path, datetime.datetime.now().time().strftime("%H%M%S") + '.xlsx')
    response = requests.get(DOWNLOAD_LINK)

    output = open(path, 'wb')
    output.write(response.content)
    output.close()
    print(path, " downloaded!")
    return path


def save_excel_data_into_database(db, table, excel_file_path):
    wb = load_workbook(excel_file_path)
    sheet = wb.active

    new_rows = []

    for row in sheet.iter_rows(min_row=50, min_col=2, max_col=8, max_row=70):
        new_rows.append((
            row[0].value,
            int(row[1].value),
            int(row[2].value),
            int(row[3].value),
            int(row[6].value))
        )
    
    # print(new_rows)

	# Connect to DB and create a cursor
    sqliteConnection = sqlite3.connect(db)

    # Get cursor
    cursor = sqliteConnection.cursor()
    query = "INSERT INTO " + table + " VALUES(?, ?, ?, ?, ?)"
    cursor.executemany(query, new_rows)
    sqliteConnection.commit()
    
    if sqliteConnection:
        sqliteConnection.close()
    print("data saved!!!")


def create_table_in_sqlite(db, table, columns):
	# Connect to DB and create a cursor
    sqliteConnection = sqlite3.connect(db)

    # Get cursor
    cursor = sqliteConnection.cursor()

    # make query phrase for createing a table called 'tsetmc' with 3 columns and then execute it
    query = "CREATE TABLE " + table +  "(" + ",".join(columns) + ")"
    cursor.execute(query)

    # Close the cursor
    cursor.close()
    
    if sqliteConnection:
        sqliteConnection.close()
    print("table created!!!")


def exist_in_option_chain(search, namad):
    # Return True if the 'search' phrase exist in 'namad'
    if search in namad:
        return True
    return False


def convert_persian_to_english_number():
    pass

def extract_strike_price(text):
    match = re.search(r'-(\d+)-', text)
    # match = re.search(r'[\u0600-\u06FF]+-(\d+)-', text)
    return match.group(1) if match else None

def extract_expiry_date(text):
    match = re.search(r'(\d{4})/(\d{2})/(\d{2})', text)
    return (match.group(1), match.group(2), match.group(3)) if match else None

