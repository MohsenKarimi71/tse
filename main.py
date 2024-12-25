import datetime
import sqlite3
import os
from openpyxl import Workbook, load_workbook
import requests

from variables import (
    DOWNLOAD_LINK,
    START_TIME,
    END_TIME,
    OPTION_CHAIN,
)
from tools import (
    download_excel_data_file,
    save_excel_data_into_database,
    create_table_in_sqlite,
    exist_in_option_chain,
    extract_strike_price,
    extract_expiry_date,
)


# create daily directory to store data files
# daily_directory_path = os.path.join('tse_excel_data', datetime.date.today().strftime("%Y%m%d"))
# os.makedirs(daily_directory_path, exist_ok=True)

counter = 0
### Download data files and Saving it into a File ###
# while(END_TIME > datetime.datetime.now().time() > START_TIME or counter == 1):
#     counter = 1
#     file_name = os.path.join(daily_directory_path)
#     print(file_name)

#     print("downloading...")
#     saved_file_path = download_excel_data_file(daily_directory_path)
    # save_excel_data_into_database(db_name, table_name, saved_file_path)


print("time finished!")



# # create_table_in_sqlite('test.db', 'mini1', ['name','count', 'volume', 'value', 'last' ])
# save_excel_data_into_database('test.db', 'mini1', os.path.join('tse_excel_data', '20241126', '112044.xlsx'))

# sqliteConnection = sqlite3.connect('test.db')
# cursor = sqliteConnection.cursor()
# result = cursor.execute("SELECT name,count,volume,value,last FROM mini1 ORDER BY value DESC")
# for row in result:
#     print(row[3],"\t ===> \t", row[0])


# Close the cursor
# cursor.close()
# if sqliteConnection:
#     sqliteConnection.close()

# wb = load_workbook("tse_excel_data/20241222/155159.xlsx")
# sheet = wb.active


# for row in sheet.iter_rows(min_row=5, max_row=280):
#     if("اختيار" in row[1].value):
#         print(row[1].value)
#         price = extract_strike_price(row[1].value)
#         print("price: ", price)

print(extract_strike_price("اختيارخ اهرم -236501-1403/10/26"))
print(extract_expiry_date("اختيارخ اهرم -236501-1403/10/26"))