from openpyxl import load_workbook
import numpy as np


### Creating Excel File Using openpyxl Library ###
# wb = Workbook()
# ws1 = wb.create_sheet("data1", 0)
# ws2 = wb.create_sheet("data0", -1)
# ws3 = wb.create_sheet("data2")

# print(wb.sheetnames)

# for sheet in wb:
#     print(sheet.title)

# ws = wb.active
# print("active sheet: ", ws.title)
# ws['A1'] = 50
# ws['B1'] = 78
# ws['C1'] = 35
# ws.cell(row=2, column=1, value="hello!")
# ws.cell(row=2, column=2, value="hello!")
# ws.cell(row=2, column=3, value="hello!")
# wb.save("test.xlsx")

### Reading Data File Using openpyxl Library ###
# wb = load_workbook("data/14030809-001.xlsx")
# sheet = wb.active

# values = []
# for i in range(4, 14):
#     values.append((sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=5).value))

# print(values)


sakook = 0
option = 0
morabehe = 0
ejare = 0
sandoogh = 0
zero_values = 0
total = 0
# wb = load_workbook(filename = 'data01.xlsx')

column_title = [
    "نماد",
    "نام",
    "تعداد",
    "حجم",
    "ارزش",
    "دیروز",
    "اولین",
    "آخرین معامله - مقدار",
    "آخرین معامله - تغییر",
    "آخرین معامله - درصد",
    "قیمت پایانی - مقدار",
    "قیمت پایانی - تغییر",
    "قیمت پایانی - درصد",
    "کمترین",
    "بیشترین",
    "EPS",
    "P/E",
    "خرید - تعداد",
    "خرید - حجم",
    "خرید - قیمت",
    "فروش - قیمت",
    "فروش - حجم",
    "فروش - تعداد"
]

column_dict = {
    "index":1,
    "name":2,
    "count":3,
    "volume":4,
    "value":5,   # final_price * volume
    "yesterday_p":6,
    "first_p":7,
    "last_p":8,
    "last_p_difference":9,
    "last_p_percent":10,
    "final_p":11,
    "final_p_difference":12,
    "final_p_percent":13,
    "min_p":14,
    "max_p":15,
    "EPS":16,
    "PE":17,
    "buy_count":18,
    "buy_volume":19,
    "buy_p":20,
    "sell_p":21,
    "sell_volume":22,
    "sell_count":23
}

percents = []
values = []



# for sheet in wb:
#     print(sheet.title)


# ws = wb.active
# for row in ws.iter_rows(min_row=4, min_col=2, max_col=10):
#     total += 1

#     if ("صكوك") in row[0].value:
#         sakook+= 1
#     elif "مرابحه" in row[0].value:
#         morabehe += 1
#     elif "اجاره" in row[0].value:
#         ejare += 1
#     elif "اختيار" in row[0].value:
#         option += 1
#     elif "صندوق" in row[0].value:
#         sandoogh += 1
#     elif row[1].value == "0":
#         zero_values += 1
#     else:
#         # percents.append([row[0].value, int(row[3].value)])
#         percents.append(float(row[8].value))
#         values.append(int(row[3].value))


# print("total: ", total)
# print("sakook: ", sakook)
# print("morabehe: ", morabehe)
# print("ejare: ", ejare)
# print("option: ", option)
# print("sandoogh: ", sandoogh)
# print("zero_values: ", zero_values)

# deleted = option + morabehe + sandoogh + ejare + sakook + zero_values
# print("deleted: ", deleted)
# print("Pure: ", total - deleted)


### getting stocks with heighest market value
# percents.sort(key=lambda ls:ls[1], reverse=True)
# for i in range(10):
#     print(percents[i][0], " >>> \t", percents[i][1])


###
avg = np.average(percents)
print("avg: ", avg)
w_avg = np.average(percents, weights=values)
print("weighted avg: ", w_avg)
